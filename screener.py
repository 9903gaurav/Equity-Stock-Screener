"""
╔══════════════════════════════════════════════════════════════════╗
║          INDIA STOCK SCREENER — NIFTY 500 WATCHLIST ENGINE       ║
║                                                                  ║
║  Steps:                                                          ║
║   1. Download price data for Nifty 500 stocks                    ║
║   2. Filter: liquidity + market cap                              ║
║   3. Rank: 12-1 momentum score                                   ║
║   4. Trend filter: price above 50 EMA & 200 EMA                  ║
║   5. Strength filter: RSI 50–70                                  ║
║   6. Volume filter: up-day volume ratio                          ║
║   7. Sector diversification check                                ║
║   8. Alert conditions: swing high breakout + volume spike        ║
║   9. Export ranked watchlist to Excel + print to console         ║
║                                                                  ║
║  Requirements:  pip install yfinance pandas numpy openpyxl       ║
║  Run:           python india_stock_screener.py                   ║
╚══════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import pandas as pd
import numpy as np
import time
import os
from datetime import datetime, timedelta

# ══════════════════════════════════════════════════════════════════
# CONFIGURATION — tweak these to change screener behaviour
# ══════════════════════════════════════════════════════════════════
CONFIG = {
    # Liquidity filters
    "min_avg_daily_volume_cr": 5,        # Minimum avg daily turnover in ₹ crores
    "min_price": 50,                     # Minimum stock price in ₹

    # Momentum
    "momentum_lookback_months": 12,      # Total lookback period
    "momentum_skip_months": 1,           # Skip last N months (avoids short-term reversal)
    "top_n_momentum": 40,                # Keep top N stocks by momentum score

    # Trend
    "ema_fast": 50,                      # Fast EMA period
    "ema_slow": 200,                     # Slow EMA period (must be above this)

    # RSI
    "rsi_period": 14,
    "rsi_min": 50,                       # RSI must be above this
    "rsi_max": 70,                       # RSI must be below this (not overbought)

    # Volume
    "vol_ratio_min": 1.0,                # Up-day volume / avg volume must exceed this
    "vol_lookback": 20,                  # Days to compute average volume

    # Breakout alert
    "swing_high_lookback": 20,           # Days to look back for recent swing high
    "breakout_vol_multiplier": 1.5,      # Volume must be this many x the average

    # Sector
    "max_per_sector": 4,                 # Max stocks per sector in final watchlist
    "final_watchlist_size": 20,          # Final number of stocks in watchlist

    # Output
    "output_file": "india_watchlist.xlsx",
}

# ══════════════════════════════════════════════════════════════════
# NIFTY 500 STOCK LIST
# A curated subset of large/mid cap NSE stocks across sectors.
# Add more tickers as needed (format: TICKER.NS)
# ══════════════════════════════════════════════════════════════════
NIFTY500_TICKERS = {
    # Banking & Finance
    "HDFCBANK.NS":   "Banking",
    "ICICIBANK.NS":  "Banking",
    "KOTAKBANK.NS":  "Banking",
    "AXISBANK.NS":   "Banking",
    "SBIN.NS":       "Banking",
    "INDUSINDBK.NS": "Banking",
    "BANDHANBNK.NS": "Banking",
    "FEDERALBNK.NS": "Banking",
    "BAJFINANCE.NS": "NBFC",
    "BAJAJFINSV.NS": "NBFC",
    "CHOLAFIN.NS":   "NBFC",
    "MUTHOOTFIN.NS": "NBFC",
    "HDFCLIFE.NS":   "Insurance",
    "SBILIFE.NS":    "Insurance",
    "ICICIPRULI.NS": "Insurance",

    # IT & Technology
    "TCS.NS":        "IT",
    "INFY.NS":       "IT",
    "WIPRO.NS":      "IT",
    "HCLTECH.NS":    "IT",
    "TECHM.NS":      "IT",
    "LTIM.NS":       "IT",
    "MPHASIS.NS":    "IT",
    "PERSISTENT.NS": "IT",
    "COFORGE.NS":    "IT",

    # FMCG & Consumer
    "HINDUNILVR.NS": "FMCG",
    "ITC.NS":        "FMCG",
    "NESTLEIND.NS":  "FMCG",
    "BRITANNIA.NS":  "FMCG",
    "DABUR.NS":      "FMCG",
    "MARICO.NS":     "FMCG",
    "COLPAL.NS":     "FMCG",
    "TATACONSUM.NS": "FMCG",

    # Auto & Auto Ancillary
    "MARUTI.NS":     "Auto",
    "TATAMOTORS.NS": "Auto",
    "M&M.NS":        "Auto",
    "BAJAJ-AUTO.NS": "Auto",
    "HEROMOTOCO.NS": "Auto",
    "EICHERMOT.NS":  "Auto",
    "BOSCHLTD.NS":   "Auto Ancillary",
    "MOTHERSON.NS":  "Auto Ancillary",
    "BALKRISIND.NS": "Auto Ancillary",

    # Pharma & Healthcare
    "SUNPHARMA.NS":  "Pharma",
    "DRREDDY.NS":    "Pharma",
    "CIPLA.NS":      "Pharma",
    "DIVISLAB.NS":   "Pharma",
    "AUROPHARMA.NS": "Pharma",
    "LUPIN.NS":      "Pharma",
    "APOLLOHOSP.NS": "Healthcare",
    "MAXHEALTH.NS":  "Healthcare",

    # Energy & Oil
    "RELIANCE.NS":   "Energy",
    "ONGC.NS":       "Energy",
    "BPCL.NS":       "Energy",
    "IOC.NS":        "Energy",
    "POWERGRID.NS":  "Utilities",
    "NTPC.NS":       "Utilities",
    "TATAPOWER.NS":  "Utilities",
    "ADANIGREEN.NS": "Utilities",
    "ADANIPORTS.NS": "Infrastructure",

    # Industrials & Capital Goods
    "LT.NS":         "Capital Goods",
    "SIEMENS.NS":    "Capital Goods",
    "ABB.NS":        "Capital Goods",
    "BHEL.NS":       "Capital Goods",
    "HAVELLS.NS":    "Capital Goods",
    "CUMMINSIND.NS": "Capital Goods",

    # Metals & Mining
    "TATASTEEL.NS":  "Metals",
    "JSWSTEEL.NS":   "Metals",
    "HINDALCO.NS":   "Metals",
    "COALINDIA.NS":  "Metals",
    "VEDL.NS":       "Metals",
    "SAIL.NS":       "Metals",
    "NMDC.NS":       "Metals",

    # Cement & Construction
    "ULTRACEMCO.NS": "Cement",
    "SHREECEM.NS":   "Cement",
    "AMBUJACEM.NS":  "Cement",
    "ACC.NS":        "Cement",
    "DALMIA.NS":     "Cement",

    # Chemicals & Specialty
    "PIDILITIND.NS": "Chemicals",
    "SRF.NS":        "Chemicals",
    "DEEPAKNTR.NS":  "Chemicals",
    "AARTI.NS":      "Chemicals",
    "TATACHEM.NS":   "Chemicals",

    # Retail & Consumer Discretionary
    "AVENUE.NS":     "Retail",
    "TRENT.NS":      "Retail",
    "NYKAA.NS":      "Retail",
    "ZOMATO.NS":     "Consumer Tech",
    "PAYTM.NS":      "Consumer Tech",

    # Telecom & Media
    "BHARTIARTL.NS": "Telecom",
    "IDEA.NS":       "Telecom",
}

# ══════════════════════════════════════════════════════════════════
# INDICATOR FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def compute_rsi(series, period=14):
    """RSI — relative strength index"""
    delta = series.diff()
    gain  = delta.clip(lower=0).rolling(period).mean()
    loss  = (-delta.clip(upper=0)).rolling(period).mean()
    rs    = gain / loss.replace(0, np.nan)
    return 100 - 100 / (1 + rs)


def compute_ema(series, span):
    """Exponential moving average"""
    return series.ewm(span=span, adjust=False).mean()


def compute_momentum_score(close, lookback_months=12, skip_months=1):
    """
    12-1 momentum: return from 12 months ago to 1 month ago.
    Skipping the last month avoids short-term mean-reversion.
    """
    trading_days = 21  # approx trading days per month
    end_idx   = len(close) - skip_months * trading_days
    start_idx = len(close) - lookback_months * trading_days

    if start_idx < 0 or end_idx <= start_idx:
        return np.nan

    start_price = close.iloc[start_idx]
    end_price   = close.iloc[end_idx]

    if start_price <= 0:
        return np.nan

    return (end_price / start_price - 1) * 100


def compute_up_volume_ratio(close, volume, lookback=20):
    """
    Ratio of average volume on up-days vs average volume on down-days.
    > 1.0 means more volume flows in on rising days = accumulation.
    """
    returns    = close.pct_change()
    up_vol     = volume[returns > 0].tail(lookback).mean()
    down_vol   = volume[returns <= 0].tail(lookback).mean()
    if down_vol == 0 or np.isnan(down_vol):
        return np.nan
    return up_vol / down_vol


def is_breakout(close, volume, swing_lookback=20, vol_multiplier=1.5):
    """
    True if today's close is above the recent swing high
    AND volume is vol_multiplier times the 20-day average.
    """
    if len(close) < swing_lookback + 2:
        return False

    recent_high  = close.iloc[-(swing_lookback+1):-1].max()
    today_close  = close.iloc[-1]
    avg_vol      = volume.iloc[-swing_lookback:-1].mean()
    today_vol    = volume.iloc[-1]

    price_break  = today_close > recent_high
    vol_confirm  = today_vol  > avg_vol * vol_multiplier

    return price_break and vol_confirm


# ══════════════════════════════════════════════════════════════════
# DATA DOWNLOAD
# ══════════════════════════════════════════════════════════════════

def download_data(tickers, period="14mo"):
    """
    Download OHLCV data for all tickers.
    Returns dict of {ticker: DataFrame}
    """
    print(f"\n{'═'*60}")
    print(f"  Downloading data for {len(tickers)} stocks...")
    print(f"{'═'*60}")

    data = {}
    failed = []

    # Download in batches of 20 to avoid rate limits
    ticker_list = list(tickers.keys())
    batch_size  = 20

    for i in range(0, len(ticker_list), batch_size):
        batch = ticker_list[i:i+batch_size]
        print(f"  Batch {i//batch_size + 1}: {', '.join([t.replace('.NS','') for t in batch])}")

        try:
            raw = yf.download(
                batch,
                period=period,
                auto_adjust=True,
                progress=False,
                threads=True,
            )

            for ticker in batch:
                try:
                    if len(batch) == 1:
                        df = raw.copy()
                    else:
                        df = raw.xs(ticker, axis=1, level=1).copy()

                    df = df.dropna(subset=["Close"])

                    if len(df) < 200:
                        failed.append(ticker)
                        continue

                    data[ticker] = df

                except Exception:
                    failed.append(ticker)

        except Exception as e:
            print(f"  Batch failed: {e}")
            failed.extend(batch)

        time.sleep(0.5)  # be polite to Yahoo Finance

    print(f"\n  Downloaded: {len(data)} stocks  |  Failed: {len(failed)}")
    if failed:
        print(f"  Failed tickers: {', '.join([t.replace('.NS','') for t in failed[:10]])}")

    return data


# ══════════════════════════════════════════════════════════════════
# SCREENING PIPELINE
# ══════════════════════════════════════════════════════════════════

def run_screener(data, tickers_sector_map):
    """
    Runs all 7 screening steps.
    Returns a ranked DataFrame with signal details.
    """

    print(f"\n{'═'*60}")
    print(f"  Running screening pipeline...")
    print(f"{'═'*60}")

    records = []

    for ticker, df in data.items():
        sector = tickers_sector_map.get(ticker, "Unknown")

        close  = df["Close"]
        volume = df["Volume"]
        high   = df["High"]

        # ── Current price ─────────────────────────────────
        price = close.iloc[-1]

        # ── Step 1: Price filter ───────────────────────────
        if price < CONFIG["min_price"]:
            continue

        # ── Step 1b: Volume filter (avg daily turnover) ───
        avg_turnover_cr = (close * volume).tail(20).mean() / 1e7  # convert to crores
        if avg_turnover_cr < CONFIG["min_avg_daily_volume_cr"]:
            continue

        # ── Step 2: Momentum score ─────────────────────────
        momentum = compute_momentum_score(
            close,
            CONFIG["momentum_lookback_months"],
            CONFIG["momentum_skip_months"],
        )
        if np.isnan(momentum):
            continue

        # ── Step 3: EMA trend ──────────────────────────────
        ema50  = compute_ema(close, CONFIG["ema_fast"]).iloc[-1]
        ema200 = compute_ema(close, CONFIG["ema_slow"]).iloc[-1]
        above_ema50  = price > ema50
        above_ema200 = price > ema200
        trend_ok = above_ema50 and above_ema200

        # ── Step 4: RSI ────────────────────────────────────
        rsi = compute_rsi(close, CONFIG["rsi_period"]).iloc[-1]
        rsi_ok = CONFIG["rsi_min"] <= rsi <= CONFIG["rsi_max"]

        # ── Step 5: Volume quality ─────────────────────────
        up_vol_ratio = compute_up_volume_ratio(
            close, volume, CONFIG["vol_lookback"]
        )
        vol_ok = (not np.isnan(up_vol_ratio)) and (up_vol_ratio >= CONFIG["vol_ratio_min"])

        # ── Step 6: Breakout alert ─────────────────────────
        breakout = is_breakout(
            close, volume,
            CONFIG["swing_high_lookback"],
            CONFIG["breakout_vol_multiplier"],
        )

        # ── Composite signal score (0–4) ───────────────────
        signal_score = sum([trend_ok, rsi_ok, vol_ok, breakout])

        # ── 52-week high/low ───────────────────────────────
        high_52w = high.tail(252).max()
        low_52w  = high.tail(252).min()
        pct_from_52w_high = (price / high_52w - 1) * 100

        records.append({
            "Ticker":           ticker.replace(".NS", ""),
            "Sector":           sector,
            "Price (₹)":        round(price, 2),
            "Momentum (%)":     round(momentum, 1),
            "RSI":              round(rsi, 1),
            "Above 50 EMA":     "✓" if above_ema50  else "✗",
            "Above 200 EMA":    "✓" if above_ema200 else "✗",
            "Up Vol Ratio":     round(up_vol_ratio, 2) if not np.isnan(up_vol_ratio) else 0,
            "Breakout Alert":   "🔔 YES" if breakout else "—",
            "Signal Score":     signal_score,
            "52w High %":       round(pct_from_52w_high, 1),
            "Avg Turnover (Cr)":round(avg_turnover_cr, 1),
            # Raw booleans for filtering
            "_trend_ok":  trend_ok,
            "_rsi_ok":    rsi_ok,
            "_vol_ok":    vol_ok,
        })

    df_all = pd.DataFrame(records)

    if df_all.empty:
        print("  No stocks passed the initial filters.")
        return df_all, pd.DataFrame()

    print(f"  Stocks after liquidity + price filter: {len(df_all)}")

    # ── Step 2: Rank by momentum, keep top N ──────────────
    df_all = df_all.sort_values("Momentum (%)", ascending=False)
    df_top = df_all.head(CONFIG["top_n_momentum"]).copy()
    print(f"  After momentum top-{CONFIG['top_n_momentum']} cut:  {len(df_top)}")

    # ── Step 3+4+5: Apply trend, RSI, volume filters ──────
    df_filtered = df_top[
        df_top["_trend_ok"] &
        df_top["_rsi_ok"]   &
        df_top["_vol_ok"]
    ].copy()
    print(f"  After trend + RSI + volume filter: {len(df_filtered)}")

    # ── Step 7: Sector cap ────────────────────────────────
    df_final = (
        df_filtered
        .groupby("Sector", group_keys=False)
        .apply(lambda g: g.head(CONFIG["max_per_sector"]))
        .sort_values("Signal Score", ascending=False)
        .head(CONFIG["final_watchlist_size"])
    )
    print(f"  After sector cap (max {CONFIG['max_per_sector']}/sector): {len(df_final)}")

    # Clean up internal columns
    display_cols = [c for c in df_final.columns if not c.startswith("_")]
    df_final = df_final[display_cols].reset_index(drop=True)
    df_final.index += 1  # start rank from 1

    return df_all[display_cols], df_final


# ══════════════════════════════════════════════════════════════════
# EXPORT TO EXCEL
# ══════════════════════════════════════════════════════════════════

def export_to_excel(df_watchlist, df_all):
    """Export results to a formatted Excel file."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter

        fname = CONFIG["output_file"]
        with pd.ExcelWriter(fname, engine="openpyxl") as writer:
            df_watchlist.to_excel(writer, sheet_name="Watchlist", index=True)
            df_all.sort_values("Momentum (%)", ascending=False).head(100).to_excel(
                writer, sheet_name="All Screened", index=False
            )

        wb = load_workbook(fname)

        # Format Watchlist sheet
        ws = wb["Watchlist"]

        header_fill  = PatternFill("solid", fgColor="1D3557")
        green_fill   = PatternFill("solid", fgColor="D4EDDA")
        alert_fill   = PatternFill("solid", fgColor="FFF3CD")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        bold_font    = Font(bold=True, size=11)
        normal_font  = Font(size=11)
        center       = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(
            bottom=Side(style="thin", color="CCCCCC")
        )

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font      = normal_font
                cell.alignment = center
                cell.border    = thin_border

            # Highlight breakout rows
            breakout_col = None
            for i, cell in enumerate(ws[1]):
                if "Breakout" in str(cell.value):
                    breakout_col = i
                    break

            if breakout_col is not None:
                if "YES" in str(row[breakout_col].value):
                    for cell in row:
                        cell.fill = alert_fill

        # Auto-size columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 25)

        ws.freeze_panes = "B2"

        wb.save(fname)
        print(f"\n  Excel saved → {fname}")

    except ImportError:
        print("  openpyxl not found — saving as CSV instead")
        df_watchlist.to_csv("india_watchlist.csv")


# ══════════════════════════════════════════════════════════════════
# CONSOLE PRINT
# ══════════════════════════════════════════════════════════════════

def print_watchlist(df):
    """Pretty-print watchlist to console."""
    print(f"\n{'═'*80}")
    print(f"  FINAL WATCHLIST  —  {datetime.now().strftime('%d %b %Y %H:%M')}")
    print(f"{'═'*80}")

    if df.empty:
        print("  No stocks made it through the screener today.")
        return

    print(f"  {'Rank':<5} {'Ticker':<14} {'Sector':<18} {'Price':>8} "
          f"{'Mom%':>7} {'RSI':>6} {'Score':>6} {'Breakout':<12}")
    print(f"  {'─'*5} {'─'*14} {'─'*18} {'─'*8} {'─'*7} {'─'*6} {'─'*6} {'─'*12}")

    for rank, row in df.iterrows():
        alert = "🔔" if "YES" in str(row.get("Breakout Alert","")) else "  "
        # print(f"  {rank:<5} {row['Ticker']:<14} {row['Sector']:<18} "
        print(f"  {rank:<5} {row['Ticker']:<14}"
              f"₹{row['Price (₹)']:>7.1f} "
              f"{row['Momentum (%)']:>+7.1f}% "
              f"{row['RSI']:>6.1f} "
              f"{row['Signal Score']:>6} "
              f"{alert} {row.get('Breakout Alert','—')}")

    print(f"\n  Sector breakdown:")
    for sector, count in df["Sector"].value_counts().items():
        bar = "█" * count
        print(f"    {sector:<20} {bar} ({count})")

    print(f"\n  Signal score explanation:")
    print(f"    4/4 = Trend ✓ + RSI ✓ + Volume ✓ + Breakout ✓  (strongest)")
    print(f"    3/4 = Three signals align  (good)")
    print(f"    Breakout alert = immediate attention needed 🔔")
    print(f"\n{'═'*80}\n")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    print(f"\n{'═'*60}")
    print(f"  INDIA STOCK SCREENER")
    print(f"  {datetime.now().strftime('%A, %d %B %Y  %H:%M IST')}")
    print(f"{'═'*60}")

    # 1. Download data
    data = download_data(NIFTY500_TICKERS)

    if not data:
        print("\nNo data downloaded. Check your internet connection.")
        return

    # 2. Run screener
    df_all, df_watchlist = run_screener(data, NIFTY500_TICKERS)

    # 3. Print to console
    print_watchlist(df_watchlist)

    # 4. Export to Excel
    if not df_watchlist.empty:
        export_to_excel(df_watchlist, df_all)

    print("  Done. Open india_watchlist.xlsx for the full report.\n")


if __name__ == "__main__":
    main()